import csv
import json
import io
from typing import Dict, Any, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import matplotlib

matplotlib.use('Agg')
import matplotlib.pyplot as plt
from io import BytesIO


class DataExporter:
    @staticmethod
    def get_currency_symbol(country: str) -> str:
        currency_map = {
            'gb': '£',
            'us': '$',
            'ca': 'CA$',
            'au': 'A$',
            'nz': 'NZ$',
            'sg': 'S$',
            'de': '€',
            'fr': '€',
            'it': '€',
            'nl': '€',
            'at': '€',
            'pl': 'zł',
            'in': '₹',
            'br': 'R$',
            'za': 'R'
        }
        return currency_map.get(country, '€')

    @staticmethod
    def to_csv(data: Dict[str, Any], country: Optional[str] = None) -> str:
        output = io.StringIO()
        currency = DataExporter.get_currency_symbol(country) if country else '€'

        output.write("SUMMARY\n")
        output.write(f"Total Jobs,{data['total_jobs']}\n")

        if data['salary_stats']['count'] > 0:
            stats = data['salary_stats']
            output.write(f"Average Salary,{currency}{stats['avg']:.2f}\n")
            output.write(f"Min Salary,{currency}{stats['min']:.2f}\n")
            output.write(f"Max Salary,{currency}{stats['max']:.2f}\n")
        output.write("\n")

        output.write("JOBS BY LOCATION\n")
        output.write("Location,Count\n")
        for location, count in data['jobs_by_location'].items():
            output.write(f"{location},{count}\n")
        output.write("\n")

        output.write("TOP SKILLS\n")
        output.write("Skill,Frequency\n")
        for skill, freq in data['top_skills']:
            output.write(f"{skill},{freq}\n")
        output.write("\n")

        if data.get('skills_by_category'):
            output.write("SKILLS BY CATEGORY\n")
            output.write("Category,Skill,Frequency\n")
            for category, skills in data['skills_by_category'].items():
                for skill, freq in skills:
                    output.write(f"{category},{skill},{freq}\n")
            output.write("\n")

        if data.get('work_type_breakdown'):
            output.write("WORK TYPE BREAKDOWN\n")
            output.write("Type,Count\n")
            for wtype, count in data['work_type_breakdown'].items():
                output.write(f"{wtype},{count}\n")
            output.write("\n")

        if data.get('experience_breakdown'):
            output.write("EXPERIENCE LEVELS\n")
            output.write("Level,Count\n")
            for level, count in data['experience_breakdown'].items():
                output.write(f"{level},{count}\n")
            output.write("\n")

        if data.get('salary_by_location'):
            output.write("SALARY BY LOCATION\n")
            output.write("Location,Average,Min,Max,Count\n")
            for loc, stats in data['salary_by_location'].items():
                output.write(f"{loc},{currency}{stats['avg']:.2f},{currency}{stats['min']:.2f},{currency}{stats['max']:.2f},{stats['count']}\n")

        return output.getvalue()

    @staticmethod
    def to_json(data: Dict[str, Any]) -> str:
        return json.dumps(data, indent=2)

    @staticmethod
    def to_xlsx(data: Dict[str, Any], country: Optional[str] = None) -> bytes:
        wb = Workbook()
        currency = DataExporter.get_currency_symbol(country) if country else '€'

        ws_summary = wb.active
        ws_summary.title = "Summary"

        header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        ws_summary['A1'] = "Metric"
        ws_summary['B1'] = "Value"
        ws_summary['A1'].fill = header_fill
        ws_summary['B1'].fill = header_fill
        ws_summary['A1'].font = header_font
        ws_summary['B1'].font = header_font

        row = 2
        ws_summary[f'A{row}'] = "Total Jobs"
        ws_summary[f'B{row}'] = data['total_jobs']

        if data['salary_stats']['count'] > 0:
            stats = data['salary_stats']
            row += 1
            ws_summary[f'A{row}'] = "Average Salary"
            ws_summary[f'B{row}'] = f"{currency}{stats['avg']:.2f}"
            row += 1
            ws_summary[f'A{row}'] = "Min Salary"
            ws_summary[f'B{row}'] = f"{currency}{stats['min']:.2f}"
            row += 1
            ws_summary[f'A{row}'] = "Max Salary"
            ws_summary[f'B{row}'] = f"{currency}{stats['max']:.2f}"

        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 20

        ws_location = wb.create_sheet("Jobs by Location")
        ws_location['A1'] = "Location"
        ws_location['B1'] = "Count"
        ws_location['A1'].fill = header_fill
        ws_location['B1'].fill = header_fill
        ws_location['A1'].font = header_font
        ws_location['B1'].font = header_font

        row = 2
        for location, count in data['jobs_by_location'].items():
            ws_location[f'A{row}'] = location
            ws_location[f'B{row}'] = count
            row += 1

        ws_location.column_dimensions['A'].width = 30
        ws_location.column_dimensions['B'].width = 15

        ws_skills = wb.create_sheet("Top Skills")
        ws_skills['A1'] = "Skill"
        ws_skills['B1'] = "Frequency"
        ws_skills['A1'].fill = header_fill
        ws_skills['B1'].fill = header_fill
        ws_skills['A1'].font = header_font
        ws_skills['B1'].font = header_font

        row = 2
        for skill, freq in data['top_skills']:
            ws_skills[f'A{row}'] = skill
            ws_skills[f'B{row}'] = freq
            row += 1

        ws_skills.column_dimensions['A'].width = 20
        ws_skills.column_dimensions['B'].width = 15

        if data.get('skills_by_category'):
            ws_cat = wb.create_sheet("Skills by Category")
            ws_cat['A1'] = "Category"
            ws_cat['B1'] = "Skill"
            ws_cat['C1'] = "Frequency"
            for col in ['A1', 'B1', 'C1']:
                ws_cat[col].fill = header_fill
                ws_cat[col].font = header_font

            row = 2
            for category, skills in data['skills_by_category'].items():
                for skill, freq in skills:
                    ws_cat[f'A{row}'] = category
                    ws_cat[f'B{row}'] = skill
                    ws_cat[f'C{row}'] = freq
                    row += 1

            ws_cat.column_dimensions['A'].width = 20
            ws_cat.column_dimensions['B'].width = 20
            ws_cat.column_dimensions['C'].width = 15

        if data.get('work_type_breakdown'):
            ws_wt = wb.create_sheet("Work Type")
            ws_wt['A1'] = "Type"
            ws_wt['B1'] = "Count"
            ws_wt['A1'].fill = header_fill
            ws_wt['B1'].fill = header_fill
            ws_wt['A1'].font = header_font
            ws_wt['B1'].font = header_font

            row = 2
            for wtype, count in data['work_type_breakdown'].items():
                ws_wt[f'A{row}'] = wtype
                ws_wt[f'B{row}'] = count
                row += 1

            ws_wt.column_dimensions['A'].width = 20
            ws_wt.column_dimensions['B'].width = 15

        if data.get('experience_breakdown'):
            ws_exp = wb.create_sheet("Experience Levels")
            ws_exp['A1'] = "Level"
            ws_exp['B1'] = "Count"
            ws_exp['A1'].fill = header_fill
            ws_exp['B1'].fill = header_fill
            ws_exp['A1'].font = header_font
            ws_exp['B1'].font = header_font

            row = 2
            for level, count in data['experience_breakdown'].items():
                ws_exp[f'A{row}'] = level
                ws_exp[f'B{row}'] = count
                row += 1

            ws_exp.column_dimensions['A'].width = 20
            ws_exp.column_dimensions['B'].width = 15

        if data.get('salary_by_location'):
            ws_sl = wb.create_sheet("Salary by Location")
            ws_sl['A1'] = "Location"
            ws_sl['B1'] = "Average"
            ws_sl['C1'] = "Min"
            ws_sl['D1'] = "Max"
            ws_sl['E1'] = "Count"
            for col in ['A1', 'B1', 'C1', 'D1', 'E1']:
                ws_sl[col].fill = header_fill
                ws_sl[col].font = header_font

            row = 2
            for loc, stats in data['salary_by_location'].items():
                ws_sl[f'A{row}'] = loc
                ws_sl[f'B{row}'] = f"{currency}{stats['avg']:.2f}"
                ws_sl[f'C{row}'] = f"{currency}{stats['min']:.2f}"
                ws_sl[f'D{row}'] = f"{currency}{stats['max']:.2f}"
                ws_sl[f'E{row}'] = stats['count']
                row += 1

            ws_sl.column_dimensions['A'].width = 25
            ws_sl.column_dimensions['B'].width = 15
            ws_sl.column_dimensions['C'].width = 15
            ws_sl.column_dimensions['D'].width = 15
            ws_sl.column_dimensions['E'].width = 10

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


class ChartExporter:

    @staticmethod
    def create_location_chart(location_data: Dict[str, int]) -> bytes:
        if not location_data or len(location_data) == 0:
            fig, ax = plt.subplots(figsize=(10, 6))
            ax.text(0.5, 0.5, 'No location data available\nTry searching with a specific city or region',
                    ha='center', va='center', fontsize=14, color='gray')
            ax.axis('off')
            output = BytesIO()
            plt.savefig(output, format='png', dpi=300, bbox_inches='tight', facecolor='white')
            plt.close(fig)
            output.seek(0)
            return output.getvalue()

        sorted_locations = sorted(location_data.items(), key=lambda x: x[1], reverse=True)[:20]

        if len(sorted_locations) == 0:
            fig, ax = plt.subplots(figsize=(10, 6))
            ax.text(0.5, 0.5, 'No location data to display',
                    ha='center', va='center', fontsize=14, color='gray')
            ax.axis('off')
            output = BytesIO()
            plt.savefig(output, format='png', dpi=300, bbox_inches='tight', facecolor='white')
            plt.close(fig)
            output.seek(0)
            return output.getvalue()

        locations = [item[0] for item in sorted_locations]
        counts = [item[1] for item in sorted_locations]

        fig, ax = plt.subplots(figsize=(14, 8))

        bars = ax.barh(locations, counts, color='#2563eb', edgecolor='#1e40af', linewidth=1.5)

        ax.set_xlabel('Number of Jobs', fontsize=14, fontweight='bold')
        ax.set_ylabel('Location', fontsize=14, fontweight='bold')

        if len(location_data) > 20:
            ax.set_title(f'Job Distribution by Location (Top 20 of {len(location_data)})',
                         fontsize=16, fontweight='bold', pad=20)
        else:
            ax.set_title('Job Distribution by Location', fontsize=16, fontweight='bold', pad=20)

        ax.invert_yaxis()

        ax.grid(axis='x', alpha=0.3, linestyle='--')
        ax.set_axisbelow(True)

        max_count = max(counts) if counts else 1
        for i, (bar, count) in enumerate(zip(bars, counts)):
            width = bar.get_width()
            ax.text(width + max_count * 0.01, bar.get_y() + bar.get_height() / 2,
                    f'{int(count)}',
                    ha='left', va='center', fontweight='bold', fontsize=10)

        plt.tight_layout()

        output = BytesIO()
        plt.savefig(output, format='png', dpi=300, bbox_inches='tight', facecolor='white')
        plt.close(fig)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def create_skills_chart(skills_data: List[List]) -> bytes:

        top_skills = skills_data[:15]
        skills = [item[0] for item in top_skills]
        frequencies = [item[1] for item in top_skills]

        fig, ax = plt.subplots(figsize=(14, 8))

        colors = [
            '#667eea',
            '#764ba2',
            '#ff6384',
            '#ff9f40',
            '#4bc0c0',
            '#9966ff',
            '#ffcd56',
            '#c9cbcf',
            '#36a2eb',
            '#667eea',
            '#764ba2',
            '#ff6384',
            '#ff9f40',
            '#4bc0c0',
            '#9966ff'
        ]

        wedges, texts = ax.pie(
            frequencies,
            labels=None,  # We'll use a legend instead
            colors=colors[:len(skills)],
            startangle=90,
            wedgeprops=dict(width=0.5, edgecolor='white', linewidth=3)
        )

        ax.legend(
            wedges,
            skills,
            title="Skills",
            loc="center left",
            bbox_to_anchor=(1, 0, 0.5, 1),
            fontsize=11,
            title_fontsize=13,
            frameon=False
        )

        ax.set_title('Top Skills in Demand', fontsize=16, fontweight='bold', pad=20)

        ax.axis('equal')

        plt.tight_layout()

        output = BytesIO()
        plt.savefig(output, format='png', dpi=300, bbox_inches='tight', facecolor='white')
        plt.close(fig)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def create_work_type_chart(work_type_data: Dict[str, int]) -> bytes:
        label_map = {
            "remote": "Remote",
            "hybrid": "Hybrid",
            "onsite": "On-site",
            "unspecified": "Unspecified",
        }
        color_map = {
            "remote": "#22c55e",
            "hybrid": "#3b82f6",
            "onsite": "#f97316",
            "unspecified": "#9ca3af",
        }

        labels = []
        sizes = []
        colors = []
        for key, count in work_type_data.items():
            if count > 0:
                labels.append(label_map.get(key, key))
                sizes.append(count)
                colors.append(color_map.get(key, "#9ca3af"))

        fig, ax = plt.subplots(figsize=(10, 8))

        if not sizes:
            ax.text(0.5, 0.5, 'No work type data available',
                    ha='center', va='center', fontsize=14, color='gray')
            ax.axis('off')
        else:
            wedges, texts, autotexts = ax.pie(
                sizes, labels=None, colors=colors, startangle=90,
                autopct='%1.1f%%',
                wedgeprops=dict(width=0.5, edgecolor='white', linewidth=3),
                pctdistance=0.75,
            )
            for t in autotexts:
                t.set_fontsize(11)
                t.set_fontweight('bold')
            ax.legend(wedges, labels, title="Work Type", loc="center left",
                      bbox_to_anchor=(1, 0, 0.5, 1), fontsize=11,
                      title_fontsize=13, frameon=False)
            ax.set_title('Work Type Breakdown', fontsize=16, fontweight='bold', pad=20)
            ax.axis('equal')

        plt.tight_layout()
        output = BytesIO()
        plt.savefig(output, format='png', dpi=300, bbox_inches='tight', facecolor='white')
        plt.close(fig)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def create_experience_chart(experience_data: Dict[str, int]) -> bytes:
        label_map = {
            "entry_level": "Entry Level",
            "mid_level": "Mid Level",
            "senior": "Senior",
            "unspecified": "Unspecified",
        }
        color_map = {
            "entry_level": "#22c55e",
            "mid_level": "#3b82f6",
            "senior": "#a855f7",
            "unspecified": "#9ca3af",
        }

        labels = [label_map.get(k, k) for k in experience_data]
        values = list(experience_data.values())
        colors = [color_map.get(k, "#9ca3af") for k in experience_data]

        fig, ax = plt.subplots(figsize=(10, 6))
        bars = ax.bar(labels, values, color=colors, edgecolor='white', linewidth=1.5, width=0.6)

        ax.set_ylabel('Number of Jobs', fontsize=13, fontweight='bold')
        ax.set_title('Experience Level Distribution', fontsize=16, fontweight='bold', pad=20)
        ax.grid(axis='y', alpha=0.3, linestyle='--')
        ax.set_axisbelow(True)

        for bar, val in zip(bars, values):
            if val > 0:
                ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + max(values) * 0.01,
                        str(val), ha='center', va='bottom', fontweight='bold', fontsize=11)

        plt.tight_layout()
        output = BytesIO()
        plt.savefig(output, format='png', dpi=300, bbox_inches='tight', facecolor='white')
        plt.close(fig)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def create_skills_category_chart(skills_by_category: Dict[str, List]) -> bytes:
        cat_totals = {}
        for cat, skills in skills_by_category.items():
            cat_totals[cat] = sum(s[1] for s in skills)

        sorted_cats = sorted(cat_totals.items(), key=lambda x: x[1], reverse=True)
        categories = [c[0] for c in sorted_cats]
        totals = [c[1] for c in sorted_cats]

        colors = ['#667eea', '#764ba2', '#ff6384', '#ff9f40',
                  '#4bc0c0', '#9966ff', '#ffcd56', '#36a2eb']

        fig, ax = plt.subplots(figsize=(12, 7))
        bars = ax.barh(categories, totals,
                       color=colors[:len(categories)],
                       edgecolor='white', linewidth=1.5)

        ax.set_xlabel('Total Mentions', fontsize=13, fontweight='bold')
        ax.set_title('Skills by Category', fontsize=16, fontweight='bold', pad=20)
        ax.invert_yaxis()
        ax.grid(axis='x', alpha=0.3, linestyle='--')
        ax.set_axisbelow(True)

        max_val = max(totals) if totals else 1
        for bar, val in zip(bars, totals):
            ax.text(bar.get_width() + max_val * 0.01,
                    bar.get_y() + bar.get_height() / 2,
                    str(val), ha='left', va='center', fontweight='bold', fontsize=10)

        plt.tight_layout()
        output = BytesIO()
        plt.savefig(output, format='png', dpi=300, bbox_inches='tight', facecolor='white')
        plt.close(fig)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def create_salary_location_chart(salary_by_location: Dict[str, Dict], currency: str = '€') -> bytes:
        sorted_locs = sorted(salary_by_location.items(),
                             key=lambda x: x[1]['avg'], reverse=True)
        locations = [item[0] for item in sorted_locs]
        avgs = [item[1]['avg'] for item in sorted_locs]

        fig, ax = plt.subplots(figsize=(12, 7))
        bars = ax.barh(locations, avgs, color='#22c55e',
                       edgecolor='#16a34a', linewidth=1.5)

        ax.set_xlabel('Average Salary', fontsize=13, fontweight='bold')
        ax.set_title('Average Salary by Location', fontsize=16, fontweight='bold', pad=20)
        ax.invert_yaxis()
        ax.grid(axis='x', alpha=0.3, linestyle='--')
        ax.set_axisbelow(True)

        max_val = max(avgs) if avgs else 1
        for bar, val in zip(bars, avgs):
            ax.text(bar.get_width() + max_val * 0.01,
                    bar.get_y() + bar.get_height() / 2,
                    f'{currency}{int(val):,}',
                    ha='left', va='center', fontweight='bold', fontsize=10)

        plt.tight_layout()
        output = BytesIO()
        plt.savefig(output, format='png', dpi=300, bbox_inches='tight', facecolor='white')
        plt.close(fig)
        output.seek(0)
        return output.getvalue()